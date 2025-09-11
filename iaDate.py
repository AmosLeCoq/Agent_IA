import fitz
import subprocess
import json
import os
import re
import time
import shutil
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from ollama import Client
from tkinter import filedialog
from tqdm import tqdm

print("Selectionner votre dossier:")
#path = input().strip()  # retire les espaces inutiles
#path = os.path.join(path, '')  # ajoute un séparateur correct à la fin
path = filedialog.askdirectory()
path=re.sub(r'/',r'\\',path)
path = os.path.join(path, '')

# On demande à PowerShell de convertir le résultat en JSON
command = f"Get-ChildItem -Path '{path}' | Select-Object Name,Length,LastWriteTime | ConvertTo-Json"
result = subprocess.run(
    ["powershell", "-Command", command],
    capture_output=True,
    text=True
)

#Création du nouveau dossier
path2 = os.path.dirname(path)
path2 = path2+"V2"
if os.path.exists(path2):
  shutil.rmtree(path2)

command3 = f"New-Item -Path '{path2}'-ItemType Directory"
result3 = subprocess.run(
    ["powershell", "-Command", command3],
    capture_output=True,
    text=True
)

#Création fichier excel
wb = Workbook()
ws = wb.active

cellNom = ws.cell(row=1, column=1)
cellNom.value = "Nom"
cellResum = ws.cell(row=1, column=2)
cellResum.value = "Résumé"
cellLien = ws.cell(row=1, column=3)
cellLien.value = "Lien"

#Fonction requete API
def requeteAPI(textAPI, modelAPI):
  try: 
    #Création de la requête à l'api
    client = Client(
      host='http://10.229.43.154:11434',
      headers={'x-some-header': 'some-value'}
    )

    if modelAPI==1:
      response = client.chat(model="mistral:latest", messages=[
        {
          'role': 'user',
          'content': textAPI,
        },
      ])

    if modelAPI==2:
      response = client.chat(model="gpt-oss:120b", messages=[
        {
          'role': 'user',
          'content': textAPI,
        },
      ])

    return response.message.content
  except Exception as e:
      print("\n\033[91mErreur API\033[0m")
      os.system("pause")
      sys.exit()

# Charger le JSON (un fichier = dict, plusieurs fichiers = list)
items = json.loads(result.stdout)

if isinstance(items, dict):
    files = [items]   # un seul fichier -> on met dans une liste
    count = 1
else:
    files = items     # plusieurs fichiers -> déjà une liste
    count = len(items)

loop = 2
loopErreur = 1


for f in tqdm(files, desc="Traitement des fichiers"):
    
    #Ouvrir le PDF
    doc = fitz.open(path+f["Name"])

    #Extraire le texte uniquement de la première page
    text = doc[0].get_text()
    nbPage=len(doc)

    #Si il y a pas de date sur la première page alors go->2->3->4
    t = 2
    while not re.search(r"(\d{1,2} \w{4,9} \d{4}|\d{1,2}[- \.]{1}\d{1,2}[- \.]{1}\d{4})", text) and nbPage >= t:
      text=doc[t-1].get_text()
      t=t+1
    # print(text)
      
    #Extraire le pdf entier
    text2 = ""
    for page in doc:
        text2 += page.get_text()

    #Edit du texte pour la création de la requête
    dateAPI = (
        "Extrait du PV ou mail :\n\n" +
        text +
        "\n\nDonne moi la date de la séance de ce pv ou mail"
        "Réponds uniquement avec la date au format YYYY-MM-DD."
        "Si tu ne trouves pas de date tu réponds uniquement /Erreur/" 
    )

    #Résumé
    resum = (
        "Extrait du PV ou mail :\n\n" +
        text2 +
        "\n\nDonne moi un résumé du en 3 phrases du pv ou mail"
    )

    #Regex pour la date
    date = requeteAPI(dateAPI, 1)
    match = re.search(r"([0-9]{4}-[0-9]{2}-[0-9]{2})", date)
    if match:
        date = match.group(1)
    else:
        date = "PasDate"+str(loopErreur)
        loopErreur = loopErreur+1

    #Déplace les pdf dans le dossier V2
    command2 = f"Copy-Item -Path '{os.path.join(path, f['Name'])}' -Destination '{path2}\\{date}.pdf' -Force"
    result2 = subprocess.run(
    ["powershell", "-Command", command2],
    capture_output=True,
    text=True
    )

    #Remplissage de l'excel
    cell1 = ws.cell(row=loop, column=1)
    cell1.value = date+".pdf"

    cell2 = ws.cell(row=loop, column=2)
    cell2.value = requeteAPI(resum, 2)

    cell3 = ws.cell(row=loop, column=3)
    cell3.value = path2+"\\"+date+".pdf"
    cell3.hyperlink = path2+"\\"+date+".pdf"
    cell3.style = "Hyperlink"

    loop=loop+1

#Création du tableau
last_row = ws.max_row
last_col = ws.max_column
ref = f"A1:{get_column_letter(last_col)}{last_row}"
tab = Table(displayName="tableau1", ref=ref)

style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

ws.add_table(tab)

wb.save(f"{path2}\\V2Excel.xlsx")

print("Installation OK")
print("Location: "+path2) 
os.system("pause")